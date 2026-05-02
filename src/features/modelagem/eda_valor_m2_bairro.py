import argparse
import hashlib
import os
import shutil
import sys
from pathlib import Path

import pandas as pd


PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.config import load_params

DEFAULT_INPUT = PROJECT_ROOT / "data" / "processed" / "itbi_features_minimal.csv"
DEFAULT_OUTPUT_ROOT = PROJECT_ROOT / "src" / "data" / "eda" / "features_modelagem"
TARGET = str(load_params().get("model", {}).get("target_column", "valor_venal_de_referencia"))
AREA = "area_do_terreno_m2"
REGION = "bairro"
FREQUENCY_QUANTILE_LIMIT = 0.99
MIN_ROWS_FOR_FREQUENCY_TRIM = 100


def dataset_version(dataset_path: Path) -> str:
    digest = hashlib.sha256()
    with dataset_path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return f"{dataset_path.stem}_{digest.hexdigest()[:12]}"


def prepare_dataset(dataset_path: Path) -> pd.DataFrame:
    if not dataset_path.exists():
        raise FileNotFoundError(f"Dataset nao encontrado: {dataset_path}")

    df = pd.read_csv(dataset_path, sep=";", low_memory=False)
    return prepare_frame(df)


def prepare_frame(df: pd.DataFrame) -> pd.DataFrame:
    required_columns = [REGION, AREA, TARGET]
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise ValueError(f"Colunas obrigatorias ausentes para EDA: {missing}")

    df = df.dropna(subset=required_columns).copy()
    df[AREA] = pd.to_numeric(df[AREA], errors="coerce")
    df[TARGET] = pd.to_numeric(df[TARGET], errors="coerce")
    df = df.dropna(subset=[AREA, TARGET]).copy()
    df = df[(df[AREA] > 0) & (df[TARGET] > 0)].copy()
    if df.empty:
        raise ValueError("Dataset sem linhas validas para calcular valor por m2.")

    df[REGION] = df[REGION].astype(str).str.strip().str.upper()
    df["valor_m2_estimado"] = df[TARGET] / df[AREA]
    return df


def bairro_summary(df: pd.DataFrame, min_samples: int = 20) -> pd.DataFrame:
    summary = (
        df.groupby(REGION)
        .agg(
            quantidade_vendas=(REGION, "size"),
            valor_m2_medio=("valor_m2_estimado", "mean"),
            valor_m2_variancia=("valor_m2_estimado", "var"),
        )
        .reset_index()
    )
    summary["valor_m2_variancia"] = summary["valor_m2_variancia"].fillna(0.0)
    summary = summary[summary["quantidade_vendas"] >= min_samples].copy()
    return summary.sort_values("valor_m2_medio", ascending=False).reset_index(drop=True)


def valor_m2_sao_paulo(df: pd.DataFrame) -> dict:
    return {
        "quantidade_vendas": int(len(df)),
        "valor_m2_medio_sao_paulo": float(df["valor_m2_estimado"].mean()),
    }


def frequency_table(series: pd.Series, bins: int) -> pd.DataFrame:
    values = series.dropna()
    if len(values) >= MIN_ROWS_FOR_FREQUENCY_TRIM:
        values = values[values <= values.quantile(FREQUENCY_QUANTILE_LIMIT)]
    intervals = pd.cut(values, bins=bins)
    counts = intervals.value_counts(sort=False)
    return pd.DataFrame(
        {
            "faixa_inicio": [float(interval.left) for interval in counts.index],
            "faixa_fim": [float(interval.right) for interval in counts.index],
            "frequencia_vendas": counts.astype(int).to_list(),
        }
    )


def export_excels(df: pd.DataFrame, summary: pd.DataFrame, output_dir: Path, bins: int) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    paths = {
        "valor_m2_por_bairro": output_dir / "valor_m2_por_bairro.xlsx",
        "top_10_bairros_maior_valor_m2": output_dir / "top_10_bairros_maior_valor_m2.xlsx",
        "bottom_30_bairros_menor_valor_m2": output_dir / "bottom_30_bairros_menor_valor_m2.xlsx",
        "frequencia_venda_por_m2": output_dir / "frequencia_venda_por_m2.xlsx",
        "frequencia_venda_por_valor": output_dir / "frequencia_venda_por_valor.xlsx",
    }

    top_10 = summary.head(10)
    bottom_30 = summary.tail(30).sort_values("valor_m2_medio")
    freq_m2 = frequency_table(df["valor_m2_estimado"], bins=bins)
    freq_valor = frequency_table(df[TARGET], bins=bins)
    sao_paulo = valor_m2_sao_paulo(df)

    _write_metric_workbook(
        path=paths["valor_m2_por_bairro"],
        data=summary,
        title="Valor medio e variancia do preco por m2 por bairro",
        category_column=REGION,
        value_column="valor_m2_medio",
        chart_rows=min(30, len(summary)),
        summary_rows=[
            ["quantidade_vendas_sao_paulo", sao_paulo["quantidade_vendas"]],
            ["valor_m2_medio_sao_paulo", sao_paulo["valor_m2_medio_sao_paulo"]],
        ],
    )
    _write_metric_workbook(
        path=paths["top_10_bairros_maior_valor_m2"],
        data=top_10,
        title="10 bairros com maior preco por m2",
        category_column=REGION,
        value_column="valor_m2_medio",
    )
    _write_metric_workbook(
        path=paths["bottom_30_bairros_menor_valor_m2"],
        data=bottom_30,
        title="30 bairros com menor preco por m2",
        category_column=REGION,
        value_column="valor_m2_medio",
        chart_rows=min(30, len(bottom_30)),
    )
    _write_frequency_workbook(
        path=paths["frequencia_venda_por_m2"],
        data=freq_m2,
        title="Frequencia de venda por valor do m2",
    )
    _write_frequency_workbook(
        path=paths["frequencia_venda_por_valor"],
        data=freq_valor,
        title="Frequencia de venda por valor de venda",
    )

    return paths


def export_images(df: pd.DataFrame, summary: pd.DataFrame, output_dir: Path, bins: int) -> dict[str, Path]:
    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    top_10 = summary.head(10)
    bottom_30 = summary.tail(30).sort_values("valor_m2_medio")
    freq_m2 = frequency_table(df["valor_m2_estimado"], bins=bins)
    freq_valor = frequency_table(df[TARGET], bins=bins)

    paths = {
        "valor_m2_por_bairro": images_dir / "valor_m2_por_bairro.png",
        "top_10_bairros_maior_valor_m2": images_dir / "top_10_bairros_maior_valor_m2.png",
        "bottom_30_bairros_menor_valor_m2": images_dir / "bottom_30_bairros_menor_valor_m2.png",
        "frequencia_venda_por_m2": images_dir / "frequencia_venda_por_m2.png",
        "frequencia_venda_por_valor": images_dir / "frequencia_venda_por_valor.png",
    }

    _save_horizontal_bar_image(
        data=summary.head(30).sort_values("valor_m2_medio"),
        category_column=REGION,
        value_column="valor_m2_medio",
        title="Valor medio do m2 por bairro",
        xlabel="Valor medio por m2",
        path=paths["valor_m2_por_bairro"],
    )
    _save_horizontal_bar_image(
        data=top_10.sort_values("valor_m2_medio"),
        category_column=REGION,
        value_column="valor_m2_medio",
        title="10 bairros com maior preco por m2",
        xlabel="Valor medio por m2",
        path=paths["top_10_bairros_maior_valor_m2"],
    )
    _save_horizontal_bar_image(
        data=bottom_30.sort_values("valor_m2_medio", ascending=False),
        category_column=REGION,
        value_column="valor_m2_medio",
        title="30 bairros com menor preco por m2",
        xlabel="Valor medio por m2",
        path=paths["bottom_30_bairros_menor_valor_m2"],
    )
    _save_frequency_image(
        data=freq_m2,
        title="Frequencia de venda por valor do m2",
        xlabel="Faixa de valor por m2",
        path=paths["frequencia_venda_por_m2"],
    )
    _save_frequency_image(
        data=freq_valor,
        title="Frequencia de venda por valor de venda",
        xlabel="Faixa de valor de venda",
        path=paths["frequencia_venda_por_valor"],
    )
    return paths


def _save_horizontal_bar_image(
    data: pd.DataFrame,
    category_column: str,
    value_column: str,
    title: str,
    xlabel: str,
    path: Path,
) -> None:
    _ensure_matplotlib_config_dir()
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(12, max(6, len(data) * 0.35)))
    ax.barh(data[category_column], data[value_column], color="#2f6f73")
    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.tick_params(axis="y", labelsize=9)
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def _save_frequency_image(data: pd.DataFrame, title: str, xlabel: str, path: Path) -> None:
    _ensure_matplotlib_config_dir()
    import matplotlib.pyplot as plt

    labels = data.apply(lambda row: f"{row['faixa_inicio']:.0f}-{row['faixa_fim']:.0f}", axis=1)
    fig, ax = plt.subplots(figsize=(14, 7))
    ax.bar(labels, data["frequencia_vendas"], color="#2f6f73")
    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel("Frequencia de vendas")
    ax.tick_params(axis="x", rotation=70, labelsize=7)
    ax.grid(axis="y", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight")
    plt.close(fig)


def _ensure_matplotlib_config_dir() -> None:
    config_dir = PROJECT_ROOT / ".test_artifacts" / "matplotlib"
    config_dir.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(config_dir))


def _write_metric_workbook(
    path: Path,
    data: pd.DataFrame,
    title: str,
    category_column: str,
    value_column: str,
    chart_rows: int | None = None,
    summary_rows: list[list[object]] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "dados"
    _write_dataframe(data_sheet, data)
    _format_data_sheet(data_sheet)

    if summary_rows:
        summary_sheet = workbook.create_sheet("resumo")
        summary_sheet.append(["metrica", "valor"])
        for row in summary_rows:
            summary_sheet.append(row)
        _format_data_sheet(summary_sheet)

    chart_sheet = workbook.create_sheet("grafico")
    chart_sheet["A1"] = title
    chart_sheet["A1"].font = Font(bold=True, size=14)
    chart_data = data.head(chart_rows or len(data))
    chart_sheet.append([category_column, value_column])
    for _, row in chart_data.iterrows():
        chart_sheet.append([row[category_column], row[value_column]])

    _format_data_sheet(chart_sheet)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = category_column
    chart.x_axis.title = value_column
    chart.height = 14
    chart.width = 24
    values = Reference(chart_sheet, min_col=2, min_row=2, max_row=len(chart_data) + 2)
    categories = Reference(chart_sheet, min_col=1, min_row=3, max_row=len(chart_data) + 2)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart_sheet.add_chart(chart, "D3")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="2F6F73")
            cell.font = Font(bold=True, color="FFFFFF")
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 38)

    workbook.save(path)


def _write_frequency_workbook(path: Path, data: pd.DataFrame, title: str) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "dados"
    data_for_excel = data.copy()
    data_for_excel["faixa"] = data_for_excel.apply(
        lambda row: f"{row['faixa_inicio']:.2f} - {row['faixa_fim']:.2f}", axis=1
    )
    _write_dataframe(data_sheet, data_for_excel[["faixa", "faixa_inicio", "faixa_fim", "frequencia_vendas"]])
    _format_data_sheet(data_sheet)

    chart_sheet = workbook.create_sheet("grafico")
    chart_sheet["A1"] = title
    chart_sheet["A1"].font = Font(bold=True, size=14)
    chart_sheet.append(["faixa", "frequencia_vendas"])
    for _, row in data_for_excel.iterrows():
        chart_sheet.append([row["faixa"], row["frequencia_vendas"]])

    _format_data_sheet(chart_sheet)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Frequencia de vendas"
    chart.x_axis.title = "Faixa"
    chart.height = 14
    chart.width = 26
    values = Reference(chart_sheet, min_col=2, min_row=2, max_row=len(data_for_excel) + 2)
    categories = Reference(chart_sheet, min_col=1, min_row=3, max_row=len(data_for_excel) + 2)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart_sheet.add_chart(chart, "D3")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="2F6F73")
            cell.font = Font(bold=True, color="FFFFFF")
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 38)

    workbook.save(path)


def _write_dataframe(sheet, data: pd.DataFrame) -> None:
    sheet.append(list(data.columns))
    for row in data.itertuples(index=False):
        sheet.append(list(row))


def _format_data_sheet(sheet) -> None:
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'


def run_eda(dataset_path: Path, output_root: Path, min_samples: int, bins: int) -> Path:
    version = dataset_version(dataset_path)
    output_dir = output_root / version
    if output_dir.exists():
        shutil.rmtree(output_dir)

    df = prepare_dataset(dataset_path)
    summary = bairro_summary(df, min_samples=min_samples)
    if summary.empty:
        raise ValueError(f"Nenhum bairro possui ao menos {min_samples} registros.")

    export_excels(df, summary, output_dir, bins=bins)
    export_images(df, summary, output_dir, bins=bins)
    return output_dir


def main() -> None:
    parser = argparse.ArgumentParser(description="EDA de valor medio por m2 por bairro.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="Dataset de features em CSV separado por ';'.")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT), help="Diretorio raiz para artefatos EDA.")
    parser.add_argument("--min-samples", type=int, default=20, help="Minimo de registros para considerar um bairro.")
    parser.add_argument("--bins", type=int, default=50, help="Quantidade de faixas nas tabelas de frequencia.")
    args = parser.parse_args()

    output_dir = run_eda(
        dataset_path=Path(args.input),
        output_root=Path(args.output_root),
        min_samples=args.min_samples,
        bins=args.bins,
    )
    print(f"EDA exportado em: {output_dir}")


if __name__ == "__main__":
    main()
