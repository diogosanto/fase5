import unittest
import shutil
from pathlib import Path

import pandas as pd

from src.features.modelagem.eda_valor_m2_bairro import (
    bairro_summary,
    dataset_version,
    export_excels,
    export_images,
    frequency_table,
    prepare_frame,
    valor_m2_sao_paulo,
)


WORKSPACE_ROOT = Path(__file__).resolve().parents[2]


class EdaValorM2BairroTests(unittest.TestCase):
    def test_prepare_frame_computes_valor_m2_and_period(self) -> None:
        prepared = prepare_frame(self._sample_frame())

        self.assertIn("valor_m2_estimado", prepared.columns)
        self.assertEqual(float(prepared.iloc[0]["valor_m2_estimado"]), 1000.0)

    def test_dataset_version_is_deterministic_for_same_file(self) -> None:
        dataset_path = Path(__file__)
        first_version = dataset_version(dataset_path)
        second_version = dataset_version(dataset_path)

        self.assertEqual(first_version, second_version)
        self.assertTrue(first_version.startswith(dataset_path.stem))

    def test_summary_ranks_bairros_and_computes_variance(self) -> None:
        frame = self._sample_frame()
        frame["valor_m2_estimado"] = frame["valor_venal_de_referencia"] / frame["area_do_terreno_m2"]
        summary = bairro_summary(frame, min_samples=1)

        self.assertEqual(summary.iloc[0]["bairro"], "B")
        self.assertIn("valor_m2_variancia", summary.columns)

    def test_valor_m2_sao_paulo_uses_all_sales(self) -> None:
        frame = prepare_frame(self._sample_frame())
        result = valor_m2_sao_paulo(frame)

        self.assertEqual(result["quantidade_vendas"], 2)
        self.assertEqual(result["valor_m2_medio_sao_paulo"], 1500.0)

    def test_frequency_table_counts_sales_by_range(self) -> None:
        table = frequency_table(pd.Series([100.0, 200.0, 300.0, 400.0]), bins=2)

        self.assertEqual(int(table["frequencia_vendas"].sum()), 4)
        self.assertEqual(len(table), 2)

    def test_export_excels_creates_five_workbooks_with_charts(self) -> None:
        from openpyxl import load_workbook

        frame = prepare_frame(self._sample_frame())
        summary = bairro_summary(frame, min_samples=1)

        temp_dir = self._artifact_dir("eda_excel_test")
        try:
            paths = export_excels(frame, summary, temp_dir, bins=2)

            self.assertEqual(len(paths), 5)
            for path in paths.values():
                self.assertTrue(path.exists())
                workbook = load_workbook(path)
                self.assertIn("dados", workbook.sheetnames)
                self.assertIn("grafico", workbook.sheetnames)
                self.assertGreaterEqual(len(workbook["grafico"]._charts), 1)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_export_images_creates_five_pngs(self) -> None:
        frame = prepare_frame(self._sample_frame())
        summary = bairro_summary(frame, min_samples=1)

        temp_dir = self._artifact_dir("eda_images_test")
        try:
            paths = export_images(frame, summary, temp_dir, bins=2)

            self.assertEqual(len(paths), 5)
            for path in paths.values():
                self.assertTrue(path.exists())
                self.assertGreater(path.stat().st_size, 0)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _sample_frame(self) -> pd.DataFrame:
        return pd.DataFrame(
            {
                "bairro": ["A", "B"],
                "area_do_terreno_m2": [100.0, 100.0],
                "valor_venal_de_referencia": [100000.0, 200000.0],
                "ano": [2024, 2024],
                "mes": [1, 1],
            }
        )

    def _artifact_dir(self, name: str) -> Path:
        path = WORKSPACE_ROOT / ".test_artifacts" / name
        shutil.rmtree(path, ignore_errors=True)
        path.mkdir(parents=True, exist_ok=True)
        return path


if __name__ == "__main__":
    unittest.main()
